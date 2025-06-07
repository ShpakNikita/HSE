import os
import sys
import io
import locale
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
import bd
from datetime import date

if sys.platform == "win32":
    locale.setlocale(locale.LC_ALL, 'rus_rus')
    os.system('chcp 65001 > nul')  

class SafeOutput:
    def __init__(self, original):
        self.original = original
    
    def write(self, text):
        try:
            if isinstance(text, str):
                self.original.write(text)
            else:
                self.original.write(text.decode('utf-8', errors='replace'))
        except:
            try:
                self.original.write(str(text).encode('ascii', errors='replace').decode('ascii'))
            except:
                self.original.write("[OUTPUT ERROR]")
    
    def flush(self):
        self.original.flush()

sys.stdout = SafeOutput(sys.stdout)
sys.stderr = SafeOutput(sys.stderr)

def print_progress(message):
    try:
        print(f" [STATUS] {message}")
    except:
        print(" [STATUS] Progress update")

current_date = date.today()
report_dir = os.path.join(os.environ['USERPROFILE'], 'Desktop', 'Еженедельные Отчеты', str(current_date))

def create_report(data, filename, sheet_name="Data", extra_data=None, extra_sheet=None):
    try:
        os.makedirs(report_dir, exist_ok=True)
        filepath = os.path.join(report_dir, filename)
        
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        if data and isinstance(data, list) and len(data) > 0:
            headers = list(data[0].keys())
            ws.append([str(h) for h in headers])
            
            for row in data:
                try:
                    ws.append([str(row.get(h, '')) for h in headers])
                except Exception as e:
                    print_progress(f"Ошибка строки: {str(e)}")
                    continue

        header_style = NamedStyle(name="header_style")
        header_style.font = Font(bold=True, color="FFFFFF")
        header_style.fill = PatternFill("solid", fgColor="4F81BD")
        header_style.border = Border(
            left=Side(style="thin"), 
            right=Side(style="thin"),
            top=Side(style="thin"), 
            bottom=Side(style="thin")
        )

        if ws.max_row > 0:
            for cell in ws[1]:
                cell.style = header_style

            for col in ws.columns:
                max_len = 0
                for cell in col:
                    try:
                        val_len = len(str(cell.value or ''))
                        max_len = max(max_len, val_len)
                    except:
                        pass
                if max_len > 0:
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

        if extra_data and extra_sheet:
            try:
                ws2 = wb.create_sheet(extra_sheet)
                if isinstance(extra_data, list) and len(extra_data) > 0:
                    headers2 = list(extra_data[0].keys())
                    ws2.append([str(h) for h in headers2])
                    
                    for row in extra_data:
                        try:
                            ws2.append([str(row.get(h, '')) for h in headers2])
                        except:
                            continue

                    for cell in ws2[1]:
                        cell.style = header_style
            except Exception as e:
                print_progress(f"Ошибка дополнительного листа: {str(e)}")

        wb.save(filepath)
        print_progress(f"Файл создан: {filename}")
        return True

    except Exception as e:
        print_progress(f"Критическая ошибка при создании {filename}: {str(e)}")
        return False

def generate_report(report_name, data_func, secondary_func=None):
    try:
        print_progress(f"Начало обработки: {report_name}")
        
        if secondary_func:
            try:
                main_data = data_func()
                extra_data = secondary_func()
                return create_report(
                    main_data,
                    f"{report_name}_{current_date}.xlsx",
                    "Основные данные",
                    extra_data,
                    "Дополнительные данные"
                )
            except Exception as e:
                print_progress(f"Ошибка получения данных: {str(e)}")
                return False
        else:
            try:
                data = data_func()
                return create_report(data, f"{report_name}_{current_date}.xlsx")
            except Exception as e:
                print_progress(f"Ошибка получения данных: {str(e)}")
                return False

    except Exception as e:
        print_progress(f"Фатальная ошибка в отчете {report_name}: {str(e)}")
        return False

def main():
    reports = [
        ("Гео на активных офферах", bd.get_geo),
        ("Активные квоты", bd.get_kvots),
        ("Актуальные закрутки", bd.get_cost, bd.get_cost_2),
        ("Сумма на офферах и продуктах", bd.get_sum, bd.get_srok),
        ("Выгрузка 1017 за 2 недели", bd.get_1017)
    ]

    success_count = 0
    for report in reports:
        if len(report) == 3:
            name, main_func, extra_func = report
            if generate_report(name, main_func, extra_func):
                success_count += 1
        else:
            name, func = report
            if generate_report(name, func):
                success_count += 1

    print_progress(f"Готово. Успешно создано: {success_count}/{len(reports)}")
    if success_count < len(reports):
        sys.exit(1)

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print_progress(f"КРИТИЧЕСКИЙ СБОЙ: {str(e)}")
        sys.exit(1)